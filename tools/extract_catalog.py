import json
from pathlib import Path

import openpyxl


SOURCE = Path(r"C:\Users\DELL\AppData\Local\Temp\HPLC Columns list - 2025 (1).xlsx")
OUTPUT = Path("data/catalog.json")


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def main():
    workbook = openpyxl.load_workbook(SOURCE, data_only=True)
    sheet = workbook.active
    keys = ["category", "packing", "phase", "particle", "pore", "diameter", "length", "code", "name"]
    rows = []

    for row_index in range(2, sheet.max_row + 1):
        values = [sheet.cell(row_index, column).value for column in range(1, sheet.max_column + 1)]
        if not any(value is not None for value in values):
            continue

        item = {key: clean(value) for key, value in zip(keys, values)}
        item["id"] = str(len(rows) + 1)
        item["search"] = " ".join(item.values()).lower()
        rows.append(item)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(rows, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {len(rows)} products to {OUTPUT}")


if __name__ == "__main__":
    main()
